import openpyxl
from openpyxl.styles import Font, Alignment

def create_marksheet_template(output_file):
    # Create a new workbook
    wb = openpyxl.Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Define subjects
    subjects = ["Subject 1", "Subject 2", "Subject 3", "Subject 4", "Subject 5", "Subject 6"]

    for subject in subjects:
        # Create a new sheet for the subject
        ws = wb.create_sheet(title=subject)

        # Set up the marksheet header
        ws['A1'] = f"Marksheet for {subject}"
        ws['A2'] = "Student Name:"
        ws['B2'] = "[Enter Name]"
        ws['A3'] = "Student ID:"
        ws['B3'] = "[Enter ID]"
        ws['A4'] = "Semester:"
        ws['B4'] = "[Enter Semester]"

        # Set up the marks table
        ws['A6'] = "Component"
        ws['B6'] = "Marks Obtained"
        ws['C6'] = "Max Marks"

        # Define components and max marks
        components = [
            ("Assignment 1", 5),
            ("Assignment 2", 5),
            ("Assignment 3", 5),
            ("Assignment 4", 5),
            ("Quiz 1", 5),
            ("Quiz 2", 5),
            ("Quiz 3", 5),
            ("Quiz 4", 5),
            ("Term Paper", 10),
            ("CP Marks/Bonus", 5),
            ("Midterms", 20),
            ("Finals", 25),
            ("Pre-Mids Total", 55),
            ("Final Total", 100)
        ]

        # Example assumption: components is a list of tuples like [("Quiz", 10), ("Assignment", 15)]
        # and ws is an object that has methods like append_row()

        # Add header
        ws.append(["Component", "Marks Obtained", "Max Marks"])

        # Populate components and max marks
        for component, max_marks in components:
            ws.append([component, "", max_marks])

        # Add total rows (adjust cell ranges as needed)
        ws.append(["Pre-Mids Total", "=SUM(B7:B10)", 55])
        ws.append(["Final Total", "=SUM(B11:B13)", 100])

        # Styling headers
        for col in ['A', 'B', 'C']:
            ws[f'{col}6'].font = Font(bold=True)
            ws[f'{col}6'].alignment = Alignment(horizontal='center')

        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15

    # Save the workbook
    wb.save(output_file)
    print(f"Marksheet template saved as {output_file}")

if __name__ == "__main__":
    output_file = "marksheet_template.xlsx"
    create_marksheet_template(output_file)