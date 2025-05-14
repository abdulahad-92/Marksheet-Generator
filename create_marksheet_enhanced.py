import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from datetime import datetime
import os

def create_marksheet_template(output_file):
    try:
        # Create a new workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Define subjects
        subjects = [
            "Statistical Inferences",
            "Intro to Macroeconomics",
            "Cal-1",
            "Philosophy",
            "Psychology",
            "IST"
        ]

        for subject in subjects:
            # Create a new sheet
            ws = wb.create_sheet(title=subject)

            # Set up header
            ws['A1'] = f"Marksheet for {subject}"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A2'] = "Student Name:"
            ws['B2'] = "[Enter Name]"
            ws['A3'] = "Student ID:"
            ws['B3'] = "[Enter ID]"
            ws['A4'] = "Semester:"
            ws['B4'] = "[Enter Semester]"
            ws['A5'] = "Last Modified:"
            ws['B5'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            # Set up marks table headers
            headers = ["Component", "Marks Obtained", "Max Marks", "Submission Status", "Remarks"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=7, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

            # Define components and max marks
            components = [
                ("Assignment 1", 5),
                ("Assignment 2", 5),
                ("Assignment 3", 5),
                ("Assignment 4", 5),
                ("Quiz 1", 5),
                ("Quiz 2", 5),
                ("Quiz 3", 5),
                ("Quiz 4", 5),
                ("Term Paper", 10),
                ("CP Marks/Bonus", 5),
                ("Midterms", 20),
                ("Finals", 25)
            ]

            # Populate components
            for row, (component, max_marks) in enumerate(components, 8):
                ws[f'A{row}'] = component
                ws[f'C{row}'] = max_marks
                ws[f'D{row}'] = "Yes" if component.startswith(("Assignment", "Term Paper")) else ""

            # Add total rows
            ws['A20'] = "Pre-Mids Total"
            ws['B20'] = "=SUM(B8:B17)"
            ws['C20'] = 55
            ws['A21'] = "Final Total"
            ws['B21'] = "=SUM(B18:B20)"
            ws['C21'] = 100

            # Add attendance tracking
            ws['A23'] = "Attendance Tracking"
            ws['A24'] = "Total Classes"
            ws['B24'] = 30  # Example value
            ws['A25'] = "Classes Attended"
            ws['B25'] = ""  # To be filled
            ws['A26'] = "Attendance %"
            ws['B26'] = "=IF(B25=\"\",0,B25/B24*100)"
            ws['B26'].number_format = '0.00%'

            # Add grade calculation
            ws['A28'] = "Grade"
            ws['B28'] = "=IF(B21>=90,\"A\",IF(B21>=80,\"B\",IF(B21>=70,\"C\",IF(B21>=60,\"D\",\"F\"))))"

            # Apply borders
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                top=Side(style='thin'), bottom=Side(style='thin'))
            for row in range(7, 22):
                for col in range(1, 6):
                    ws.cell(row=row, column=col).border = thin_border
            for row in range(24, 27):
                for col in range(1, 3):
                    ws.cell(row=row, column=col).border = thin_border
            ws['B28'].border = thin_border

            # Data validation for marks
            for row in range(8, 20):
                max_marks = ws[f'C{row}'].value
                dv = DataValidation(type="decimal", operator="between", formula1=0, formula2=max_marks)
                dv.add(f'B{row}')
                ws.add_data_validation(dv)

            # Data validation for submission status
            dv_status = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
            dv_status.add('D8:D13')
            ws.add_data_validation(dv_status)

            # Conditional formatting for grades
            ws.conditional_formatting.add('B28',
                CellIsRule(operator='equal', formula=['"A"'], fill=PatternFill(start_color="90EE90", fill_type="solid")))
            ws.conditional_formatting.add('B28',
                CellIsRule(operator='equal', formula=['"B"'], fill=PatternFill(start_color="ADD8E6", fill_type="solid")))
            ws.conditional_formatting.add('B28',
                CellIsRule(operator='equal', formula=['"C"'], fill=PatternFill(start_color="FFFFE0", fill_type="solid")))
            ws.conditional_formatting.add('B28',
                CellIsRule(operator='equal', formula=['"D"'], fill=PatternFill(start_color="FFA07A", fill_type="solid")))
            ws.conditional_formatting.add('B28',
                CellIsRule(operator='equal', formula=['"F"'], fill=PatternFill(start_color="FF4040", fill_type="solid")))

            # Adjust column widths
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 10
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 25

            # Freeze panes
            ws.freeze_panes = 'A8'

        # Save the workbook
        wb.save(output_file)
        print(f"Marksheet template saved as {output_file}")

    except Exception as e:
        print(f"Error creating marksheet: {e}")

if __name__ == "__main__":
    output_file = "marksheet_template_enhanced.xlsx"
    create_marksheet_template(output_file)