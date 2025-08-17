import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from datetime import datetime
import os
import tkinter as tk
from tkinter import messagebox, ttk
import logging
import string
import re

# Configure logging for accountability
logging.basicConfig(
    filename='marksheet_creation_log.txt',
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

def sanitize_filename(text):
    """Sanitize input for valid filename."""
    # Remove invalid characters, replace spaces with underscores
    text = re.sub(r'[<>:"/\\|?*]', '', text.strip())
    text = text.replace(' ', '_')
    # Ensure non-empty and valid length
    return text[:50] or "Unknown"

def create_marksheet_template(output_file, subjects, student_name, student_id, semester):
    """Generate Excel marksheet with given subjects and student details."""
    try:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        for subject in subjects:
            ws = wb.create_sheet(title=subject)
            ws['A1'] = f"Marksheet for {subject}"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A2'] = "Student Name:"
            ws['B2'] = student_name or "[Enter Name]"
            ws['A3'] = "Student ID:"
            ws['B3'] = student_id or "[Enter ID]"
            ws['A4'] = "Semester:"
            ws['B4'] = semester or "[Enter Semester]"
            ws['A5'] = "Last Modified:"
            ws['B5'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            headers = ["Component", "Marks Obtained", "Max Marks", "Submission Status", "Remarks"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=7, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

            components = [
                ("Assignment 1", 5), ("Assignment 2", 5), ("Assignment 3", 5), ("Assignment 4", 5),
                ("Quiz 1", 5), ("Quiz 2", 5), ("Quiz 3", 5), ("Quiz 4", 5),
                ("Term Paper", 10), ("CP Marks/Bonus", 5), ("Midterms", 20), ("Finals", 25)
            ]

            for row, (component, max_marks) in enumerate(components, 8):
                ws[f'A{row}'] = component
                ws[f'C{row}'] = max_marks
                ws[f'D{row}'] = "Yes" if component.startswith(("Assignment", "Term Paper")) else ""

            ws['A20'] = "Pre-Mids Total"
            ws['B20'] = "=SUM(B8:B17)"
            ws['C20'] = 55
            ws['A21'] = "Final Total"
            ws['B21'] = "=SUM(B18:B20)"
            ws['C21'] = 100

            ws['A23'] = "Attendance Tracking"
            ws['A24'] = "Total Classes"
            ws['B24'] = 30
            ws['A25'] = "Classes Attended"
            ws['B25'] = ""
            ws['A26'] = "Attendance %"
            ws['B26'] = "=IF(B25=\"\",0,B25/B24*100)"
            ws['B26'].number_format = '0.00%'

            ws['A28'] = "Grade"
            ws['B28'] = "=IF(B21>=90,\"A\",IF(B21>=80,\"B\",IF(B21>=70,\"C\",IF(B21>=60,\"D\",\"F\"))))"

            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                top=Side(style='thin'), bottom=Side(style='thin'))
            for row in range(7, 22):
                for col in range(1, 6):
                    ws.cell(row=row, column=col).border = thin_border
            for row in range(24, 27):
                for col in range(1, 3):
                    ws.cell(row=row, column=col).border = thin_border
            ws['B28'].border = thin_border

            for row in range(8, 20):
                max_marks = ws[f'C{row}'].value
                dv = DataValidation(type="decimal", operator="between", formula1=0, formula2=max_marks)
                dv.add(f'B{row}')
                ws.add_data_validation(dv)

            dv_status = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
            dv_status.add('D8:D13')
            ws.add_data_validation(dv_status)

            ws.conditional_formatting.add('B28',
                CellIsRule(operator='equal', formula=['"A"'], fill=PatternFill(start_color="90EE90", fill_type="solid")))
            ws.conditional_formatting.add('B28',
                CellIsRule(operator='equal', formula=['"B"'], fill=PatternFill(start_color="ADD8E6", fill_type="solid")))
            ws.conditional_formatting.add('B28',
                CellIsRule(operator='equal', formula=['"C"'], fill=PatternFill(start_color="FFFFE0", fill_type="solid")))
            ws.conditional_formatting.add('B28',
                CellIsRule(operator='equal', formula=['"D"'], fill=PatternFill(start_color="FFA07A", fill_type="solid")))
            ws.conditional_formatting.add('B28',
                CellIsRule(operator='equal', formula=['"F"'], fill=PatternFill(start_color="FF4040", fill_type="solid")))

            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 10
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 25
            ws.freeze_panes = 'A8'

        wb.save(output_file)
        logging.info(f"Marksheet template saved as {output_file}")
        return True
    except Exception as e:
        logging.error(f"Error creating marksheet: {e}")
        return str(e)

class MarksheetApp:
    """Modern Tkinter GUI for creating marksheet with dynamic course input."""
    def __init__(self, root):
        self.root = root
        self.root.title("Marksheet Generator")
        self.root.geometry("700x600")
        self.root.configure(bg="#F0F4F8")  # Light blue-gray background

        # Styling
        self.style = ttk.Style()
        self.style.configure("TButton", font=("Helvetica", 10), padding=10)
        self.style.configure("TLabel", font=("Helvetica", 12), background="#F0F4F8")
        self.style.configure("TEntry", font=("Helvetica", 10))

        # Lists to store course entries
        self.course_entries = []

        # Main frame
        self.main_frame = tk.Frame(root, bg="#F0F4F8")
        self.main_frame.pack(padx=20, pady=20, fill="both", expand=True)

        # Header
        tk.Label(self.main_frame, text="Marksheet Generator", font=("Helvetica", 18, "bold"), bg="#F0F4F8", fg="#1E3A8A").pack(pady=10)

        # Student details frame
        self.student_frame = tk.Frame(self.main_frame, bg="#FFFFFF", bd=2, relief="groove")
        self.student_frame.pack(pady=10, padx=10, fill="x")

        tk.Label(self.student_frame, text="Student Name:", font=("Helvetica", 12), bg="#FFFFFF").grid(row=0, column=0, padx=10, pady=5, sticky="e")
        self.name_entry = ttk.Entry(self.student_frame, width=40)
        self.name_entry.grid(row=0, column=1, padx=10, pady=5)

        tk.Label(self.student_frame, text="Student ID:", font=("Helvetica", 12), bg="#FFFFFF").grid(row=1, column=0, padx=10, pady=5, sticky="e")
        self.id_entry = ttk.Entry(self.student_frame, width=40)
        self.id_entry.grid(row=1, column=1, padx=10, pady=5)

        tk.Label(self.student_frame, text="Semester:", font=("Helvetica", 12), bg="#FFFFFF").grid(row=2, column=0, padx=10, pady=5, sticky="e")
        self.semester_entry = ttk.Entry(self.student_frame, width=40)
        self.semester_entry.grid(row=2, column=1, padx=10, pady=5)

        # Courses frame with scrollbar
        tk.Label(self.main_frame, text="Courses:", font=("Helvetica", 12, "bold"), bg="#F0F4F8", fg="#1E3A8A").pack(pady=10)
        self.course_canvas = tk.Canvas(self.main_frame, bg="#FFFFFF", bd=2, relief="groove", height=200)
        self.course_scrollbar = ttk.Scrollbar(self.main_frame, orient="vertical", command=self.course_canvas.yview)
        self.course_frame = tk.Frame(self.course_canvas, bg="#FFFFFF")
        self.course_canvas.configure(yscrollcommand=self.course_scrollbar.set)
        self.course_canvas.pack(side="left", fill="both", expand=True, padx=10)
        self.course_scrollbar.pack(side="right", fill="y")
        self.course_canvas.create_window((0, 0), window=self.course_frame, anchor="nw")
        self.course_frame.bind("<Configure>", lambda e: self.course_canvas.configure(scrollregion=self.course_canvas.bbox("all")))

        # Buttons
        self.button_frame = tk.Frame(self.main_frame, bg="#F0F4F8")
        self.button_frame.pack(pady=10)
        ttk.Button(self.button_frame, text="Add Course", command=self.add_course, style="Accent.TButton").pack(side="left", padx=5)
        ttk.Button(self.button_frame, text="Generate Marksheet", command=self.generate_marksheet, style="Accent.TButton").pack(side="left", padx=5)

        # Status label
        self.status_label = tk.Label(self.main_frame, text="", font=("Helvetica", 10), bg="#F0F4F8", fg="green")
        self.status_label.pack(pady=10)

        # Style for buttons
        self.style.configure("Accent.TButton", background="#3B82F6", foreground="black")
        self.style.map("Accent.TButton", background=[("active", "#2563EB")])

        # Add initial course entry
        self.add_course()

    def add_course(self):
        """Add a new course entry field."""
        row = len(self.course_entries)
        entry = ttk.Entry(self.course_frame, width=50)
        entry.grid(row=row, column=0, padx=5, pady=5)
        remove_btn = ttk.Button(self.course_frame, text="Remove", command=lambda: self.remove_course(entry, remove_btn))
        remove_btn.grid(row=row, column=1, padx=5, pady=5)
        self.course_entries.append((entry, remove_btn))
        logging.info("Added new course entry field")
        self.course_canvas.configure(scrollregion=self.course_canvas.bbox("all"))

    def remove_course(self, entry, button):
        """Remove a course entry field."""
        if len(self.course_entries) > 1:
            self.course_entries.remove((entry, button))
            entry.destroy()
            button.destroy()
            logging.info("Removed course entry field")
            self.course_canvas.configure(scrollregion=self.course_canvas.bbox("all"))
        else:
            messagebox.showwarning("Warning", "At least one course is required!")
            logging.warning("Attempted to remove last course entry")

    def generate_marksheet(self):
        """Generate the Excel marksheet based on user input."""
        student_name = self.name_entry.get().strip()
        student_id = self.id_entry.get().strip()
        semester = self.semester_entry.get().strip()
        subjects = [entry.get().strip() for entry, _ in self.course_entries if entry.get().strip()]

        if not subjects:
            self.status_label.config(text="Error: Please enter at least one course!", fg="red")
            messagebox.showerror("Error", "Please enter at least one course!")
            logging.error("No courses provided")
            return

        if not student_name or not semester:
            self.status_label.config(text="Error: Student name and semester are required!", fg="red")
            messagebox.showerror("Error", "Student name and semester are required for file naming!")
            logging.error("Missing student name or semester")
            return

        # Generate filename using student name and semester
        filename = f"marksheet_{sanitize_filename(student_name)}_{sanitize_filename(semester)}.xlsx"
        output_file = os.path.join(os.path.dirname(__file__), filename)

        result = create_marksheet_template(output_file, subjects, student_name, student_id, semester)
        if result is True:
            self.status_label.config(text=f"Marksheet saved as {filename}", fg="green")
            messagebox.showinfo("Success", f"Marksheet generated successfully: {filename}")
            logging.info(f"Marksheet generated for {len(subjects)} courses")
        else:
            self.status_label.config(text=f"Error: {result}", fg="red")
            messagebox.showerror("Error", f"Failed to generate marksheet: {result}")

if __name__ == "__main__":
    root = tk.Tk()
    app = MarksheetApp(root)
    root.mainloop()